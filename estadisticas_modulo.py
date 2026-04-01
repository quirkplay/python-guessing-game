"""
=========================== MODULO DE GESTION DE ESTADISTICAS ===========================
                        Maneja Excel, pandas y todos los rankings 
"""

import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd

# ===================================== GESTIÓN DE LA HOJA DE EXCEL =====================================

def inicializar_excel():
    """Crea el archivo Excel de estadísticas si no existe"""
    ruta_fichero = "estadisticas.xlsx"
    
    if os.path.exists(ruta_fichero) == False:
        print("Creando archivo de estadísticas...")
        
        # Crear nuevo libro
        documento = Workbook()
        hoja = documento.active
        hoja.title = "Estadísticas"
        
        # Crear lista de encabezados
        encabezados = ["Nombre", "Resultado", "Intentos", "Dificultad", "Modo", "Rol", "Fecha", "Hora"]
        
        # Añadir encabezados en la primera fila
        hoja.append(encabezados)
        
        # Guardar
        documento.save(ruta_fichero)
        print("Archivo creado correctamente.")



def guardar_resultado(nombre, gano, intentos, dificultad, modo, rol):
    """Guarda el resultado de una partida"""
    ruta_fichero = "estadisticas.xlsx"
    
    # Asegurar que existe el archivo
    inicializar_excel()
    
    # Cargar el archivo
    documento = load_workbook(ruta_fichero)
    hoja = documento.active
    
    # Determinar si es victoria o derrota
    if gano == True:
        resultado = "Victoria"
    else:
        resultado = "Derrota"
    
    # Obtener fecha y hora actuales
    ahora = datetime.now()
    fecha = ahora.strftime("%d/%m/%Y")  # Formato: 12/11/2025
    hora = ahora.strftime("%H:%M:%S")   # Formato: 14:35:22
    
    # Crear lista con todos los datos
    nueva_fila = [nombre, resultado, intentos, dificultad, modo, rol, fecha, hora]
    
    # Añadir la fila al Excel
    hoja.append(nueva_fila)
    
    # Guardar cambios
    documento.save(ruta_fichero)
    
    print(f'\n✅ Resultado de {nombre} guardado correctamente.')



# ===================================== FUNCIONES DE ESTADISTICAS GENERALES =====================================


def mostrar_todas_estadisticas():
    """Muestra todas las estadísticas guardadas"""
    ruta_fichero = "estadisticas.xlsx"
    
    df = pd.read_excel(ruta_fichero)
    
    print("\n" + "="*100)
    print("TODAS LAS ESTADÍSTICAS")
    print("="*100)
    print()
    
    if len(df) == 0:
        print("⚠️  No hay partidas registradas")
    else:
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', 100) 
        pd.set_option('display.max_colwidth', 15)
        
        print(df.to_string(index=False))
        
        print("\n" + "="*100)
        
        partidas_solitario = len(df[df['Modo'] == 'Solitario'])
        partidas_dos_jugadores = len(df[df['Modo'] == '2 Jugadores']) // 2
        total_partidas = partidas_solitario + partidas_dos_jugadores
        
        print(f"Total de partidas: {total_partidas}")
        print(f"  • Modo Solitario: {partidas_solitario}")
        print(f"  • Modo 2 Jugadores: {partidas_dos_jugadores}")
        print(f"Total de registros: {len(df)}")
    
    input("\nPresiona Enter para continuar...")



# ===================================== FUNCIONES DE BUSQUEDA POR JUGADOR =====================================



def buscar_mis_estadisticas():
    """Busca y muestra estadísticas de un jugador"""
    ruta_fichero = "estadisticas.xlsx"
    
    datos_encontrados = False
    
    while datos_encontrados == False:
        print("\n--- BUSCAR MIS ESTADÍSTICAS ---")
        print("\nNota: El buscador no distingue mayúsculas y minúsculas para facilitar tu búsqueda.")
        nombre_buscar = input("Introduce tu nombre (o presiona Enter para salir): ").strip().lower()
        
        if nombre_buscar == "":
            return
        
        # Cargar datos del jugador
        datos_jugador = obtener_datos_jugador(nombre_buscar, ruta_fichero)
        
        # Verificar si existen partidas
        if datos_jugador["total_partidas"] == 0:
            print(f"\n⚠️  No se encontraron partidas para '{nombre_buscar}'")
            print("Verifica que el nombre esté escrito correctamente")
        else:
            datos_encontrados = True
    
    # Mostrar estadísticas
    mostrar_estadisticas_jugador(nombre_buscar, datos_jugador)
    
    input("\nPresiona Enter para continuar...")



def obtener_datos_jugador(nombre_buscar, ruta_fichero):
    """Analiza todas las partidas de un jugador y devuelve sus estadísticas"""
    
    # Cargar Excel
    documento = load_workbook(ruta_fichero)
    hoja = documento.active
    
    # Inicializar contadores
    datos = {
        "total_partidas": 0,
        "solitario_total": 0,
        "solitario_victorias": 0,
        "dos_jug_adivinador_total": 0,
        "dos_jug_adivinador_victorias": 0,
        "dos_jug_maestro_total": 0,
        "dos_jug_maestro_victorias": 0,
        "ultima_fecha": None,
        "ultima_hora": None
    }
    
    # Recorrer todas las filas que contienen datos menos la de los títulos.
    for fila_num in range(2, hoja.max_row + 1):
        nombre_fila = hoja.cell(row=fila_num, column=1).value
        
        # Comparar nombres (ambos en minúsculas)
        if nombre_fila and nombre_fila.lower() == nombre_buscar.lower():
            # Obtengo el dato de cada fila
            resultado = hoja.cell(row=fila_num, column=2).value
            modo = hoja.cell(row=fila_num, column=5).value
            rol = hoja.cell(row=fila_num, column=6).value
            fecha = hoja.cell(row=fila_num, column=7).value
            hora = hoja.cell(row=fila_num, column=8).value
            
            # Contar partida total
            datos["total_partidas"] = datos["total_partidas"] + 1
            
            # Clasificar por modo y rol
            if modo == "Solitario":
                datos["solitario_total"] = datos["solitario_total"] + 1
                if resultado == "Victoria":
                    datos["solitario_victorias"] = datos["solitario_victorias"] + 1
            
            elif modo == "2 Jugadores":
                if rol == "Adivinador":
                    datos["dos_jug_adivinador_total"] = datos["dos_jug_adivinador_total"] + 1
                    if resultado == "Victoria":
                        datos["dos_jug_adivinador_victorias"] = datos["dos_jug_adivinador_victorias"] + 1
                
                elif rol == "Desafiante":
                    datos["dos_jug_maestro_total"] = datos["dos_jug_maestro_total"] + 1
                    if resultado == "Victoria":
                        datos["dos_jug_maestro_victorias"] = datos["dos_jug_maestro_victorias"] + 1
            
            # Actualizar última partida
            datos["ultima_fecha"] = fecha
            datos["ultima_hora"] = hora
    
    return datos



def mostrar_estadisticas_jugador(nombre, datos):
    """Muestra las estadísticas de un jugador especifico"""
    
    print("\n" + "="*60)
    print(f"  ESTADÍSTICAS DE {nombre.upper()}")
    print("="*60)
    
    # Partidas totales
    print(f"\n📊 PARTIDAS TOTALES: {datos['total_partidas']}")
    
    # Modo Solitario
    print(f"\n🎮 MODO SOLITARIO")
    print(f"   Partidas jugadas: {datos['solitario_total']}")
    if datos['solitario_total'] > 0:
        porcentaje_sol = (datos['solitario_victorias'] / datos['solitario_total']) * 100
        print(f"   Victorias: {datos['solitario_victorias']}")
        print(f"   Porcentaje de victoria: {porcentaje_sol:.1f}%")
    else:
        print(f"   No has jugado en este modo")
    
    # Modo 2 Jugadores - Adivinador
    print(f"\n🎯 MODO 2 JUGADORES (Adivinador)")
    print(f"   Partidas jugadas: {datos['dos_jug_adivinador_total']}")
    if datos['dos_jug_adivinador_total'] > 0:
        porcentaje_adiv = (datos['dos_jug_adivinador_victorias'] / datos['dos_jug_adivinador_total']) * 100
        print(f"   Victorias: {datos['dos_jug_adivinador_victorias']}")
        print(f"   Porcentaje de victoria: {porcentaje_adiv:.1f}%")
    else:
        print(f"   No has jugado en este modo")
    
    # Modo 2 Jugadores - Maestro
    print(f"\n🎩 MODO 2 JUGADORES (Maestro)")
    print(f"   Partidas jugadas: {datos['dos_jug_maestro_total']}")
    if datos['dos_jug_maestro_total'] > 0:
        porcentaje_maestro = (datos['dos_jug_maestro_victorias'] / datos['dos_jug_maestro_total']) * 100
        print(f"   Victorias: {datos['dos_jug_maestro_victorias']}")
        print(f"   Porcentaje de victoria: {porcentaje_maestro:.1f}%")
    else:
        print(f"   No has jugado en este modo")
    
    # Última partida
    print(f"\n⏰ ÚLTIMA PARTIDA")
    print(f"   Fecha: {datos['ultima_fecha']}")
    print(f"   Hora: {datos['ultima_hora']}")
    
    print("\n" + "="*60)



# ===================================== FUNCIONES DE RANKINGS =====================================



def menu_rankings():
    """Menú de rankings con modos"""
    
    continuar = True
    
    while continuar:
        print("\n" + "="*50)
        print(" "*15 + "RANKINGS")
        print("="*50)
        print("\n--- FILTRAR POR MODO ---")
        print("\n1. Modo Solitario")
        print("2. Modo 2 Jugadores")
        print("3. Todos los modos")
        print("4. Volver a Estadisticas")
        print("="*50)
        
        opcion = input("\nElige una opción: ")
        
        if opcion.isdigit():
            opcion = int(opcion)
            
            if 1 <= opcion <= 4:
                if opcion == 1:
                    seleccionar_dificultad_ranking("Solitario")
                elif opcion == 2:
                    seleccionar_dificultad_ranking("2 Jugadores")
                elif opcion == 3:
                    seleccionar_dificultad_ranking(None)  # Todos los modos
                elif opcion == 4:
                    continuar = False
            else:
                print("ERROR: Debes elegir entre 1 y 4")
        else:
            print("ERROR: Debes introducir un número válido")


            
def seleccionar_dificultad_ranking(modo_filtro):
    """Permite seleccionar dificultad y muestra el menú de tops"""
    
    continuar = True
    
    while continuar:
        print("\n--- FILTRAR POR DIFICULTAD ---")
        
        if modo_filtro:
            print(f"Modo: {modo_filtro}")
        else:
            print("Modo: Todos")
        
        print("\n1. Fácil")
        print("2. Medio")
        print("3. Difícil")
        print("4. Todas las dificultades")
        print("5. Volver a Seleccionar Modo")
        
        opcion = input("\nElige: ")
        
        if opcion.isdigit():
            opcion = int(opcion)
            
            if 1 <= opcion <= 5:
                if opcion == 1:
                    mostrar_menu_tops(modo_filtro, "Fácil")
                elif opcion == 2:
                    mostrar_menu_tops(modo_filtro, "Medio")
                elif opcion == 3:
                    mostrar_menu_tops(modo_filtro, "Difícil")
                elif opcion == 4:
                    mostrar_menu_tops(modo_filtro, None)  # Todas las dificultades
                elif opcion == 5:
                    continuar = False
            else:
                print("ERROR: Debes elegir entre 1 y 5")
        else:
            print("ERROR: Debes introducir un número válido")


            
def mostrar_menu_tops(modo_filtro, dificultad_filtro):
    """Muestra el menú de tops según los filtros seleccionados"""
    
    continuar = True
    
    while continuar:
        print("\n" + "="*50)
        print(" "*10 + "TOP 5 - RANKINGS")
        print("="*50)
        
        if modo_filtro:
            print(f"Modo: {modo_filtro}")
        else:
            print("Modo: Todos")
            
        if dificultad_filtro:
            print(f"Dificultad: {dificultad_filtro}")
        else:
            print("Dificultad: Todas")
        
        print("="*50)
        print("1. Top 5 victorias con menos intentos")
        print("2. Top 5 jugadores con más victorias")
        print("3. Top 5 rachas de victorias")
        print("4. Volver a Seleccionar Dificultad")
        print("="*50)
        
        opcion = input("\nElige: ")
        
        if opcion.isdigit():
            opcion = int(opcion)
            
            if 1 <= opcion <= 4:
                if opcion == 1:
                    top_menos_intentos(modo_filtro, dificultad_filtro)
                elif opcion == 2:
                    top_mas_victorias(modo_filtro, dificultad_filtro)
                elif opcion == 3:
                    top_rachas_victorias(modo_filtro, dificultad_filtro)
                elif opcion == 4:
                    continuar = False
            else:
                print("ERROR: Debes elegir entre 1 y 4")
        else:
            print("ERROR: Debes introducir un número válido")



def top_menos_intentos(modo_filtro, dificultad_filtro):
    """Muestra el top 5 de victorias con menos intentos"""
    
    ruta_fichero = "estadisticas.xlsx"
    df = pd.read_excel(ruta_fichero)
    
    # Filtrar solo victorias
    df = df[df['Resultado'] == 'Victoria']
    
    # Aplicar filtros de modo y de dificultad
    if modo_filtro:
        df = df[df['Modo'] == modo_filtro]
    
    if dificultad_filtro:
        df = df[df['Dificultad'] == dificultad_filtro]
    
    # Verificar si hay datos
    if len(df) == 0:
        print("\n⚠️  No hay victorias con estos filtros")
        input("\nPresiona Enter para continuar...")
        return
    
    # Ordenar por intentos de menor a mayor y tomar los primeros 5
    df_top = df.nsmallest(5, 'Intentos')
    
    # Mostrar resultado
    print("\n" + "="*100)
    print("🏆 TOP 5 - VICTORIAS CON MENOS INTENTOS")
    print("="*100)
    
    if modo_filtro:
        print(f"Modo: {modo_filtro}")
    else:
        print("Modo: Todos")
        
    if dificultad_filtro:
        print(f"Dificultad: {dificultad_filtro}")
    else:
        print("Dificultad: Todas")
    
    print("\n")
    print(df_top[['Nombre', 'Intentos', 'Dificultad', 'Modo', 'Fecha', 'Hora']].to_string(index=False))
    print("\n" + "="*100)
    
    input("\nPresiona Enter para continuar...")



def top_mas_victorias(modo_filtro, dificultad_filtro):
    """Muestra el top 5 de jugadores con más victorias"""
    
    ruta_fichero = "estadisticas.xlsx"
    df = pd.read_excel(ruta_fichero)
    
    # Filtrar solo victorias
    df = df[df['Resultado'] == 'Victoria']
    
    # Aplicar filtros de modo y dificultad
    if modo_filtro:
        df = df[df['Modo'] == modo_filtro]
        
    if dificultad_filtro:
        df = df[df['Dificultad'] == dificultad_filtro]
    
    # Verificar si hay datos
    if len(df) == 0:
        print("\n⚠️  No hay victorias con estos filtros")
        input("\nPresiona Enter para continuar...")
        return
    
    # Contar victorias por jugador
    victorias_por_jugador = df['Nombre'].value_counts().head(5) #value_counts me cuenta cuantas veces se repite ese nombre y head(5) solo toma los 5 primeros
    
    # Mostrar resultado
    print("\n" + "="*60)
    print("🏆 TOP 5 - JUGADORES CON MÁS VICTORIAS")
    print("="*60)
    
    if modo_filtro:
        print(f"Modo: {modo_filtro}")
    else:
        print("Modo: Todos")
        
    if dificultad_filtro:
        print(f"Dificultad: {dificultad_filtro}")
    else:
        print("Dificultad: Todas")
    
    print("\n")
    
    posicion = 1
    
    for item in victorias_por_jugador.items():
        nombre = item[0]
        victorias = item[1]
        print(f"{posicion}. {nombre}: {victorias} victorias")
        posicion = posicion + 1
    
    print("\n" + "="*60)
    
    input("\nPresiona Enter para continuar...")



def top_rachas_victorias(modo_filtro, dificultad_filtro):
    """Muestra el top 5 de rachas de victorias consecutivas"""
    
    ruta_fichero = "estadisticas.xlsx"
    df = pd.read_excel(ruta_fichero)
    
    # Aplicar filtros de modo y de dificultad
    if modo_filtro:
        df = df[df['Modo'] == modo_filtro]
    
    if dificultad_filtro:
        df = df[df['Dificultad'] == dificultad_filtro]
    
    # Verificar si hay datos
    if len(df) == 0:
        print("\n⚠️  No hay partidas con estos filtros")
        input("\nPresiona Enter para continuar...")
        return
    
    # Calcular rachas por jugador
    rachas = calcular_rachas_por_jugador(df)
    
    # Verificar si hay rachas
    if len(rachas) == 0:
        print("\n⚠️  No hay rachas de victorias")
        input("\nPresiona Enter para continuar...")
        return
    
    # Ordenar por longitud de racha y tomar top 5
    rachas_top5 = sorted(rachas, key=obtener_racha, reverse=True)[:5]
    
    # Mostrar resultado
    print("\n" + "="*60)
    print("🏆 TOP 5 - RACHAS DE VICTORIAS CONSECUTIVAS")
    print("="*60)
    
    if modo_filtro:
        print(f"Modo: {modo_filtro}")
    else:
        print("Modo: Todos")
        
    if dificultad_filtro:
        print(f"Dificultad: {dificultad_filtro}")
    else:
        print("Dificultad: Todas")
    
    print("\n")

    posicion = 1
    
    for tupla in rachas_top5:
        nombre = tupla[0]
        racha = tupla[1]
        print(f"{posicion}. {nombre}: {racha} victorias consecutivas")
        posicion = posicion + 1
    
    print("\n" + "="*60)
    
    input("\nPresiona Enter para continuar...")



def calcular_rachas_por_jugador(df):
    """Calcula la racha más larga de victorias para cada jugador"""
    rachas = []
    
    # Agrupar por jugador
    for nombre in df['Nombre'].unique():
        # Obtener todas las partidas de ese jugador
        partidas_jugador = df[df['Nombre'] == nombre].copy()
        
        # Calcular la racha máxima
        racha_actual = 0
        racha_maxima = 0
        
        for resultado in partidas_jugador['Resultado']:
            if resultado == 'Victoria':
                racha_actual = racha_actual + 1
                if racha_actual > racha_maxima:
                    racha_maxima = racha_actual
            else:
                racha_actual = 0
        
        # Solo agregar si tiene al menos una victoria
        if racha_maxima > 0:
            rachas.append((nombre, racha_maxima))
    
    return rachas


    
def obtener_racha(tupla):
    """Devuelve el número de racha para ordenar"""
    return tupla[1]
    #Funcion auxiliar para usar el sorted en la funcion top_rachas_victorias